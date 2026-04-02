import os
import json
import threading
import requests
from datetime import datetime, timedelta
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from flask import Flask, request, jsonify, send_file
from apscheduler.schedulers.background import BackgroundScheduler
import openpyxl
from openpyxl import load_workbook
import pytz

app = Flask(__name__)
client = WebClient(token=os.environ.get("SLACK_BOT_TOKEN"))
NPT = pytz.timezone("Asia/Kathmandu")

# Timing
STANDUP_HOUR = int(os.environ.get("STANDUP_HOUR", "7"))
STANDUP_MINUTE = int(os.environ.get("STANDUP_MINUTE", "20"))
STANDUP_WINDOW_MINUTES = int(os.environ.get("STANDUP_WINDOW", "30"))
RENDER_URL = os.environ.get("RENDER_URL", "https://dohoro-standup.onrender.com")
ADMIN_SLACK_IDS = [i.strip() for i in os.environ.get("ADMIN_SLACK_ID", "").split(",") if i.strip()]

# Teams
TEAMS = {
    "dev":  {"members": os.environ.get("DEV_MEMBERS", ""),  "channel": os.environ.get("DEV_CHANNEL", "")},
    "qa":   {"members": os.environ.get("QA_MEMBERS", ""),   "channel": os.environ.get("QA_CHANNEL", "")},
    "uiux": {"members": os.environ.get("UIUX_MEMBERS", ""), "channel": os.environ.get("UIUX_CHANNEL", "")},
}

QUESTIONS = [
    "👋 Good morning! Time for your daily standup!\n\n*Question 1 of 3:* ✅ What did you complete yesterday?",
    "*Question 2 of 3:* 🔨 What will you work on today?",
    "*Question 3 of 3:* 🚧 Anything blocking your progress? (Type 'none' if nothing)"
]

SESSIONS_FILE = "/tmp/sessions.json"
SUBMITTED_FILE = "/tmp/submitted_today.json"
EXCEL_LOCK = threading.Lock()
SESSIONS_LOCK = threading.Lock()

# ─── Session helpers ───────────────────────────────────────────────────────────

def load_sessions():
    try:
        with open(SESSIONS_FILE, "r") as f:
            return json.load(f)
    except:
        return {}

def save_sessions(sessions):
    try:
        with open(SESSIONS_FILE, "w") as f:
            json.dump(sessions, f)
    except Exception as e:
        print(f"Error saving sessions: {e}")

def load_submitted():
    try:
        with open(SUBMITTED_FILE, "r") as f:
            return json.load(f)
    except:
        return {}

def save_submitted(data):
    try:
        with open(SUBMITTED_FILE, "w") as f:
            json.dump(data, f)
    except Exception as e:
        print(f"Error saving submitted: {e}")

def mark_submitted(user_id, message_ts, channel_id, team_name, team_channel):
    with SESSIONS_LOCK:
        submitted = load_submitted()
        submitted[user_id] = {
            "message_ts": message_ts,
            "channel_id": channel_id,
            "team_name": team_name,
            "team_channel": team_channel
        }
        save_submitted(submitted)

def get_submitted(user_id):
    with SESSIONS_LOCK:
        return load_submitted().get(user_id)

def clear_submitted():
    save_submitted({})

def get_session(user_id):
    with SESSIONS_LOCK:
        return load_sessions().get(user_id)

def set_session(user_id, data):
    with SESSIONS_LOCK:
        sessions = load_sessions()
        sessions[user_id] = data
        save_sessions(sessions)

def delete_session(user_id):
    with SESSIONS_LOCK:
        sessions = load_sessions()
        sessions.pop(user_id, None)
        save_sessions(sessions)

def get_all_sessions():
    with SESSIONS_LOCK:
        return load_sessions()

# ─── Helpers ───────────────────────────────────────────────────────────────────

def keep_alive():
    try:
        requests.get(RENDER_URL, timeout=10)
        print(f"✅ Keep alive at {datetime.now(NPT).strftime('%H:%M NPT')}")
    except Exception as e:
        print(f"Keep alive failed: {e}")

def get_close_time():
    close_min = STANDUP_MINUTE + STANDUP_WINDOW_MINUTES
    close_hour = STANDUP_HOUR + close_min // 60
    close_min = close_min % 60
    return close_hour, close_min

def is_standup_open():
    now = datetime.now(NPT)
    open_time = now.replace(hour=STANDUP_HOUR, minute=STANDUP_MINUTE, second=0, microsecond=0)
    close_time = open_time + timedelta(minutes=STANDUP_WINDOW_MINUTES)
    return open_time <= now <= close_time

def get_all_team_members():
    all_members = {}
    for team_name, team_config in TEAMS.items():
        members_str = team_config["members"]
        if not members_str.strip():
            continue
        for m in members_str.split(","):
            name = m.strip().lower()
            if name:
                if name in all_members:
                    print(f"⚠️ Warning: {name} is in multiple teams! Using first team: {all_members[name][0]}")
                else:
                    all_members[name] = (team_name, team_config["channel"])
    return all_members

def notify_admin(message):
    if not ADMIN_SLACK_IDS:
        print(f"ADMIN ALERT: {message}")
        return
    for admin_id in ADMIN_SLACK_IDS:
        try:
            dm = client.conversations_open(users=admin_id)
            client.chat_postMessage(
                channel=dm["channel"]["id"],
                text=f"🤖 *Dohoro Standup Alert*\n{message}"
            )
        except SlackApiError as e:
            print(f"Could not notify admin {admin_id}: {e}")

# ─── Excel ─────────────────────────────────────────────────────────────────────

def get_excel_filepath(team_name):
    now = datetime.now(NPT)
    return f"/tmp/standup_{team_name}_{now.strftime('%B-%Y')}.xlsx"

def save_to_excel(user_name, team_name, answers, status="Submitted"):
    filepath = get_excel_filepath(team_name)
    now = datetime.now(NPT)
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M")
    with EXCEL_LOCK:
        try:
            wb = load_workbook(filepath)
            ws = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = now.strftime("%B %Y")
            ws.append(["Date", "Time", "Name", "Team", "Yesterday", "Today", "Blockers", "Status"])
        if status == "Did not submit":
            ws.append([date_str, time_str, user_name, team_name.upper(), "-", "-", "-", "Did not submit"])
        else:
            blockers = answers[2] if answers[2].lower() != "none" else "-"
            ws.append([date_str, time_str, user_name, team_name.upper(), answers[0], answers[1], blockers, status])
        wb.save(filepath)

# ─── Slack posting ─────────────────────────────────────────────────────────────

def get_channel_id(channel_name):
    if not channel_name:
        return None
    try:
        result = client.conversations_list(types="public_channel,private_channel")
        for ch in result["channels"]:
            if ch["name"] == channel_name.strip():
                return ch["id"]
        notify_admin(f"⚠️ Channel `#{channel_name}` not found! Please check the channel name in Render settings.")
    except SlackApiError as e:
        print(f"Error getting channel {channel_name}: {e}")
    return None

def update_channel_message(user_name, user_id, team_name, channel_name, answers, message_ts, channel_id):
    blockers = answers[2] if answers[2].lower() != "none" else "-"
    team_badge = get_team_badge(team_name)
    team_color = get_team_color(team_name)
    now = datetime.now(NPT)
    date_str = now.strftime("%B %d, %Y · %I:%M %p")
    try:
        client.chat_update(
            channel=channel_id,
            ts=message_ts,
            text=f"{user_name} updated standup - DOHORO-STANDUP ☕",
            attachments=[
                {
                    "color": team_color,
                    "blocks": [
                        {
                            "type": "section",
                            "text": {"type": "mrkdwn", "text": f"*<@{user_id}> submitted standup* ☕   `{team_badge}`   `✏️ Edited`"}
                        },
                        {"type": "divider"},
                        {
                            "type": "section",
                            "text": {"type": "mrkdwn", "text": f"*✅ What did you complete yesterday?*\n{answers[0]}"}
                        },
                        {
                            "type": "section",
                            "text": {"type": "mrkdwn", "text": f"*💻 What will you work on today?*\n{answers[1]}"}
                        },
                        {
                            "type": "section",
                            "text": {"type": "mrkdwn", "text": f"*🚧 Anything blocking your progress?*\n{blockers}"}
                        },
                        {
                            "type": "context",
                            "elements": [{"type": "mrkdwn", "text": f"DOHORO-STANDUP · {date_str} _(edited)_"}]
                        }
                    ]
                }
            ]
        )
        print(f"✅ Updated message for {user_name} ({team_name})")
    except SlackApiError as e:
        print(f"Error updating message: {e}")

def get_team_color(team_name):
    colors = {
        "dev": "#22c55e",
        "qa": "#ef4444",
        "uiux": "#7c3aed"
    }
    return colors.get(team_name.lower(), "#378ADD")

def get_team_badge(team_name):
    badges = {
        "dev": "💻 DEV",
        "qa": "🧪 QA",
        "uiux": "🎨 UIUX"
    }
    return badges.get(team_name.lower(), team_name.upper())

def post_to_channel(user_name, user_id, team_name, channel_name, answers, late=False):
    if not channel_name:
        print(f"No channel configured for {team_name} team!")
        notify_admin(f"⚠️ No channel configured for `{team_name}` team! Please set `{team_name.upper()}_CHANNEL` in Render.")
        return
    channel_id = get_channel_id(channel_name)
    if not channel_id:
        return
    blockers = answers[2] if answers[2].lower() != "none" else "-"
    team_badge = get_team_badge(team_name)
    team_color = get_team_color(team_name)
    status_text = "⏰ Submitted late" if late else "✅ Submitted on time"
    from datetime import datetime
    now = datetime.now(NPT)
    date_str = now.strftime("%B %d, %Y · %I:%M %p")

    try:
        response = client.chat_postMessage(
            channel=channel_id,
            text=f"{user_name} submitted standup - DOHORO-STANDUP ☕",
            attachments=[
                {
                    "color": team_color,
                    "blocks": [
                        {
                            "type": "section",
                            "text": {
                                "type": "mrkdwn",
                                "text": f"*<@{user_id}> submitted standup* ☕   `{team_badge}`   `{status_text}`"
                            }
                        },
                        {
                            "type": "divider"
                        },
                        {
                            "type": "section",
                            "text": {
                                "type": "mrkdwn",
                                "text": f"*✅ What did you complete yesterday?*\n{answers[0]}"
                            }
                        },
                        {
                            "type": "section",
                            "text": {
                                "type": "mrkdwn",
                                "text": f"*💻 What will you work on today?*\n{answers[1]}"
                            }
                        },
                        {
                            "type": "section",
                            "text": {
                                "type": "mrkdwn",
                                "text": f"*🚧 Anything blocking your progress?*\n{blockers}"
                            }
                        },
                        {
                            "type": "context",
                            "elements": [
                                {
                                    "type": "mrkdwn",
                                    "text": f"DOHORO-STANDUP · {date_str}"
                                }
                            ]
                        }
                    ]
                }
            ]
        )
        message_ts = response.get("ts")
        print(f"✅ Posted {user_name} ({team_name}) to #{channel_name}{' [LATE]' if late else ''}")
        return message_ts, channel_id
    except SlackApiError as e:
        print(f"Error posting to #{channel_name}: {e}")
        notify_admin(f"⚠️ Failed to post {user_name}'s standup to `#{channel_name}`: {e}")
    return None, None

def post_did_not_submit(user_name, user_id, channel_name):
    if not channel_name:
        return
    channel_id = get_channel_id(channel_name)
    if not channel_id:
        return
    try:
        client.chat_postMessage(
            channel=channel_id,
            text=f"⚠️ <@{user_id}> did not submit standup today."
        )
    except SlackApiError as e:
        print(f"Error posting missed standup: {e}")

# ─── Standup flow ──────────────────────────────────────────────────────────────

def close_standup():
    close_hour, close_min = get_close_time()
    print(f"Closing standup at {close_hour}:{close_min:02d} NPT")
    sessions = get_all_sessions()
    for user_id, session in sessions.items():
        user_name = session["name"]
        dm_channel = session["channel"]
        team_name = session.get("team", "unknown")
        team_channel = session.get("team_channel", "")
        try:
            client.chat_postMessage(
                channel=dm_channel,
                text=f"⏰ Standup is now closed ({close_hour}:{close_min:02d} deadline reached). Your response was not recorded. Please submit on time tomorrow!"
            )
        except SlackApiError:
            pass
        post_did_not_submit(user_name, user_id, team_channel)
        save_to_excel(user_name, team_name, [], status="Did not submit")
    save_sessions({})
    print("Standup closed!")

def send_standup_prompts():
    print(f"Sending standup prompts at {datetime.now(NPT).strftime('%H:%M NPT')}")
    save_sessions({})
    clear_submitted()
    member_map = get_all_team_members()
    if not member_map:
        print("No team members configured!")
        notify_admin("⚠️ No team members configured! Please set DEV_MEMBERS, QA_MEMBERS, or UIUX_MEMBERS in Render.")
        return
    try:
        result = client.users_list()
        sent_count = 0
        for user in result["members"]:
            if user.get("is_bot") or user.get("deleted") or user.get("is_app_user"):
                continue
            if user["id"] == "USLACKBOT":
                continue
            user_real_name = user.get("real_name", user.get("name", "")).lower()
            if user_real_name not in member_map:
                continue
            team_name, team_channel = member_map[user_real_name]
            user_id = user["id"]
            try:
                dm = client.conversations_open(users=user_id)
                dm_channel = dm["channel"]["id"]
                set_session(user_id, {
                    "step": 0,
                    "answers": [],
                    "channel": dm_channel,
                    "name": user.get("real_name", user.get("name", "Team member")),
                    "team": team_name,
                    "team_channel": team_channel
                })
                client.chat_postMessage(channel=dm_channel, text=QUESTIONS[0])
                sent_count += 1
                print(f"✅ Sent to: {user.get('real_name')} ({team_name})")
            except SlackApiError as e:
                print(f"Error DMing {user_id}: {e}")
        print(f"Standup sent to {sent_count} members!")
    except SlackApiError as e:
        print(f"Error fetching users: {e}")
        notify_admin(f"⚠️ Error fetching Slack users: {e}")

# ─── Slack events ──────────────────────────────────────────────────────────────

@app.route("/slack/events", methods=["POST"])
def slack_events():
    # Ignore Slack retry attempts to prevent duplicate questions
    if request.headers.get("X-Slack-Retry-Num"):
        return jsonify({"status": "ok"})

    data = request.json
    if not data:
        return jsonify({"status": "ok"})

    if "challenge" in data:
        return jsonify({"challenge": data["challenge"]})
    if "event" in data:
        event = data["event"]
        if event.get("type") == "message" and not event.get("bot_id"):
            user_id = event.get("user")
            text = event.get("text", "").strip()
            channel = event.get("channel")
            session = get_session(user_id)
            print(f"Event from {user_id}, session exists: {session is not None}")

            # Edit command detection
            if not session and text.lower().strip() in ["edit", "edit standup", "change", "update"]:
                submitted = get_submitted(user_id)
                if submitted:
                    set_session(user_id, {
                        "step": 0,
                        "answers": [],
                        "channel": channel,
                        "name": user_id,
                        "team": submitted["team_name"],
                        "team_channel": submitted["team_channel"],
                        "late": False,
                        "editing": True,
                        "edit_ts": submitted["message_ts"],
                        "edit_channel_id": submitted["channel_id"]
                    })
                    try:
                        client.chat_postMessage(
                            channel=channel,
                            text="✏️ No problem! Let us update your standup. " + QUESTIONS[0]
                        )
                    except SlackApiError:
                        pass
                    return jsonify({"status": "ok"})
                else:
                    try:
                        client.chat_postMessage(
                            channel=channel,
                            text="❌ You have not submitted a standup today yet. Please submit first!"
                        )
                    except SlackApiError:
                        pass
                    return jsonify({"status": "ok"})

            # Option A: Reopen session for late submission if no active session
            if not session:
                member_map = get_all_team_members()
                try:
                    user_info = client.users_info(user=user_id)
                    user_real_name = user_info["user"].get("real_name", "").lower()
                    if user_real_name in member_map:
                        team_name, team_channel = member_map[user_real_name]
                        # Check if they already submitted today by checking if session was completed
                        # Start fresh late session
                        set_session(user_id, {
                            "step": 0,
                            "answers": [],
                            "channel": channel,
                            "name": user_info["user"].get("real_name", "Team member"),
                            "team": team_name,
                            "team_channel": team_channel,
                            "late": not is_standup_open()
                        })
                        session = get_session(user_id)
                        try:
                            client.chat_postMessage(
                                channel=channel,
                                text="⏰ Standup window has closed but we will still accept your submission as *late*. " + QUESTIONS[0]
                            )
                        except SlackApiError:
                            pass
                        return jsonify({"status": "ok"})
                except SlackApiError:
                    pass

            if not user_id:
                return jsonify({"status": "ok"})

            if session and session["channel"] == channel:
                session["answers"].append(text)
                session["step"] += 1
                current_step = session["step"]
                answers = list(session["answers"])
                user_name = session["name"]
                team_name = session.get("team", "unknown")
                team_channel = session.get("team_channel", "")
                if current_step < len(QUESTIONS):
                    set_session(user_id, session)
                    try:
                        client.chat_postMessage(channel=channel, text=QUESTIONS[current_step])
                        print(f"✅ Sent Q{current_step+1} to {user_name}")
                    except SlackApiError as e:
                        print(f"Error sending Q{current_step+1}: {e}")
                else:
                    delete_session(user_id)
                    late = session.get("late", False) or not is_standup_open()
                    status = "Submitted late" if late else "Submitted"
                    try:
                        if late:
                            client.chat_postMessage(
                                channel=channel,
                                text="✅ *Thank you! Your standup has been submitted!* 🚀\n_(Note: This was submitted after the deadline but has been recorded as late.)_"
                            )
                        else:
                            client.chat_postMessage(
                                channel=channel,
                                text="✅ *Thank you! Your standup has been submitted!* 🚀\nHave a productive day! 💪"
                            )
                    except SlackApiError:
                        pass
                    post_to_channel(user_name, user_id, team_name, team_channel, answers, late=late)
                    save_to_excel(user_name, team_name, answers, status=status)
                    print(f"✅ Standup complete for {user_name} ({team_name}) [{status}]")
    return jsonify({"status": "ok"})

# ─── Routes ────────────────────────────────────────────────────────────────────

@app.route("/", methods=["GET"])
def home():
    close_hour, close_min = get_close_time()
    now = datetime.now(NPT)
    result = f"<h3>Dohoro Standup Bot ☕</h3>"
    result += f"<b>Current Nepal time:</b> {now.strftime('%H:%M NPT')}<br>"
    result += f"<b>Standup opens:</b> {STANDUP_HOUR}:{STANDUP_MINUTE:02d} NPT<br>"
    result += f"<b>Standup closes:</b> {close_hour}:{close_min:02d} NPT<br><br>"
    result += f"<b>Status:</b> {'🟢 Open' if is_standup_open() else '🔴 Closed'}<br><br>"
    for team_name, team_config in TEAMS.items():
        members = team_config["members"]
        channel = team_config["channel"]
        if members:
            count = len([m for m in members.split(",") if m.strip()])
            result += f"<b>{team_name.upper()} team</b> ({count} members) → #{channel}<br>"
    result += f"<br><a href='/members'>View all members</a> | "
    result += f"<a href='/download'>Download reports</a>"
    return result

@app.route("/members", methods=["GET"])
def list_members():
    result = "<h3>Standup Teams</h3>"
    for team_name, team_config in TEAMS.items():
        members_str = team_config["members"]
        channel = team_config["channel"]
        if members_str.strip():
            result += f"<b>{team_name.upper()} team</b> → #{channel}<br>"
            for m in members_str.split(","):
                result += f"&nbsp;&nbsp;• {m.strip()}<br>"
            result += "<br>"
    return result

@app.route("/download", methods=["GET"])
def download_all():
    result = "<h3>Download Standup Reports</h3>"
    now = datetime.now(NPT)
    for team_name in TEAMS.keys():
        result += f"<a href='/download/{team_name}'>{team_name.upper()} team — {now.strftime('%B %Y')}</a><br>"
    return result

@app.route("/download/<team>", methods=["GET"])
def download_excel(team):
    if team not in TEAMS:
        return f"Team '{team}' not found! Valid teams: dev, qa, uiux", 404
    filepath = get_excel_filepath(team)
    try:
        now = datetime.now(NPT)
        return send_file(
            filepath,
            as_attachment=True,
            download_name=f"standup_{team}_{now.strftime('%B-%Y')}.xlsx"
        )
    except FileNotFoundError:
        return f"No Excel file found for {team.upper()} team yet. Data will appear after first standup submission!", 404

@app.route("/trigger", methods=["GET"])
def trigger():
    thread = threading.Thread(target=send_standup_prompts)
    thread.start()
    return "Standup triggered! Check Slack DMs 🚀"

@app.route("/close", methods=["GET"])
def manual_close():
    thread = threading.Thread(target=close_standup)
    thread.start()
    return "Standup closed manually!"

# ─── Scheduler ─────────────────────────────────────────────────────────────────

def start_scheduler():
    close_hour, close_min = get_close_time()
    scheduler = BackgroundScheduler(timezone=NPT)
    scheduler.add_job(
        send_standup_prompts, "cron",
        day_of_week="sun,mon,tue,wed,thu,fri",
        hour=STANDUP_HOUR, minute=STANDUP_MINUTE
    )
    scheduler.add_job(
        close_standup, "cron",
        day_of_week="sun,mon,tue,wed,thu,fri",
        hour=close_hour, minute=close_min
    )
    scheduler.add_job(keep_alive, "interval", minutes=5)
    scheduler.start()
    print(f"✅ Scheduler started!")
    print(f"✅ Standup opens: {STANDUP_HOUR}:{STANDUP_MINUTE:02d} NPT")
    print(f"✅ Standup closes: {close_hour}:{close_min:02d} NPT")
    print(f"✅ Keep alive: every 5 minutes")

def initialize():
    if not os.environ.get("SCHEDULER_STARTED"):
        os.environ["SCHEDULER_STARTED"] = "1"
        start_scheduler()

initialize()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 3000)))
